# src/routes/testing.py
"""
测试平台 API 路由

提供基于模型版本的推理评估功能
"""

import asyncio
import queue
from datetime import datetime
from typing import Annotated, Any, Dict, Optional, cast

from fastapi import (
    APIRouter,
    Depends,
    HTTPException,
    Query,
    WebSocket,
    WebSocketDisconnect,
    status,
)
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from src.config import get_local_now
from src.models import (
    BatteryUnit,
    ModelVersion,
    TestJob,
    TestJobBattery,
    User,
    get_db,
)
from src.routes.auth import get_current_user
from src.tasks.testing_worker import start_test_job

router = APIRouter()


# --- WebSocket for Real-time Progress ---
class ConnectionManager:
    """WebSocket 连接管理器（支持跨线程消息推送）"""

    def __init__(self):
        self.active_connections: dict[int, list[WebSocket]] = {}
        self.message_queue: queue.Queue = queue.Queue()
        self._broadcast_task: Optional[asyncio.Task] = None

    async def connect(self, job_id: int, websocket: WebSocket):
        """连接 WebSocket"""
        await websocket.accept()
        if job_id not in self.active_connections:
            self.active_connections[job_id] = []
        self.active_connections[job_id].append(websocket)

        # 启动消息广播任务
        if self._broadcast_task is None or self._broadcast_task.done():
            self._broadcast_task = asyncio.create_task(self._message_broadcaster())

    def disconnect(self, job_id: int, websocket: WebSocket):
        """断开 WebSocket"""
        if job_id in self.active_connections:
            try:
                self.active_connections[job_id].remove(websocket)
            except ValueError:
                pass
            if not self.active_connections[job_id]:
                del self.active_connections[job_id]

    def push_message(self, job_id: int, message: Dict[str, Any]):
        """线程安全的消息推送（供Worker线程调用）"""
        try:
            self.message_queue.put_nowait((job_id, message))
        except queue.Full:
            pass

    async def send_message(self, job_id: int, message: dict):
        """向指定任务的所有连接发送消息（异步方法）"""
        if job_id in self.active_connections:
            dead_connections = []
            for connection in self.active_connections[job_id]:
                try:
                    await connection.send_json(message)
                except Exception:
                    dead_connections.append(connection)

            for conn in dead_connections:
                try:
                    self.active_connections[job_id].remove(conn)
                except ValueError:
                    pass

    async def _message_broadcaster(self):
        """后台任务：从队列中读取消息并广播"""
        while True:
            try:
                try:
                    job_id, message = self.message_queue.get_nowait()
                    await self.send_message(job_id, message)
                except queue.Empty:
                    await asyncio.sleep(0.05)
            except Exception as e:
                print(f"消息广播错误: {e}")
                await asyncio.sleep(0.1)


manager = ConnectionManager()


# --- Schemas ---
class CreateTestJobRequest(BaseModel):
    """创建测试任务请求"""

    model_version_id: int
    dataset_id: int
    target: str = Field(..., pattern="^(RUL|PCL|BOTH)$")
    battery_ids: list[int] = Field(..., min_length=1)
    horizon: int = Field(default=1, ge=1)


class TestJobResponse(BaseModel):
    """测试任务响应"""

    id: int
    user_id: int
    model_version_id: int
    dataset_id: int
    target: str
    horizon: int
    status: str
    created_at: datetime
    started_at: Optional[datetime]
    finished_at: Optional[datetime]

    class Config:
        from_attributes = True


class TestJobDetailResponse(BaseModel):
    """测试任务详情响应"""

    job: TestJobResponse
    model_info: dict
    batteries: list[dict]


# --- API Endpoints ---


@router.post(
    "/jobs", response_model=TestJobResponse, status_code=status.HTTP_201_CREATED
)
async def create_test_job(
    request: CreateTestJobRequest,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Session = Depends(get_db),
):
    """
    创建测试任务

    - 验证模型版本所有权
    - 验证数据集访问权限
    - 验证电池ID有效性
    - 创建测试任务并启动后台Worker
    """
    # 1. 验证模型版本所有权
    model_version = (
        db.query(ModelVersion)
        .filter(
            ModelVersion.id == request.model_version_id,
            ModelVersion.user_id == current_user.id,
        )
        .first()
    )

    if not model_version:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="模型版本不存在或无权访问"
        )

    # 2. 验证数据集访问权限
    from src.models import Dataset

    dataset = db.query(Dataset).filter(Dataset.id == request.dataset_id).first()

    if not dataset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="数据集不存在"
        )

    if bool(dataset.source_type == "UPLOAD") and bool(
        dataset.owner_user_id != current_user.id
    ):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN, detail="无权访问此数据集"
        )

    # 3. 验证电池ID有效性
    batteries = (
        db.query(BatteryUnit)
        .filter(
            BatteryUnit.id.in_(request.battery_ids),
            BatteryUnit.dataset_id == request.dataset_id,
        )
        .all()
    )

    if len(batteries) != len(request.battery_ids):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="部分电池ID无效"
        )

    # 4. 创建测试任务
    test_job = TestJob(
        user_id=current_user.id,
        model_version_id=request.model_version_id,
        dataset_id=request.dataset_id,
        target=request.target,
        horizon=request.horizon,
        status="PENDING",
        created_at=get_local_now(),
    )
    db.add(test_job)
    db.flush()

    job_id: int = test_job.id  # type: ignore[assignment]

    # 5. 创建电池关联
    for battery_id in request.battery_ids:
        test_battery = TestJobBattery(test_job_id=job_id, battery_id=battery_id)
        db.add(test_battery)

    db.commit()
    db.refresh(test_job)

    # 6. 启动后台测试任务
    start_test_job(job_id=job_id)

    return test_job


@router.get("/jobs/{job_id}", response_model=TestJobDetailResponse)
async def get_test_job(
    job_id: int,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Session = Depends(get_db),
):
    """
    获取测试任务详情

    - 任务基本信息
    - 模型版本信息
    - 关联的电池列表
    """
    job = (
        db.query(TestJob)
        .filter(TestJob.id == job_id, TestJob.user_id == current_user.id)
        .first()
    )

    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="测试任务不存在"
        )

    # 查询模型信息
    model_version = (
        db.query(ModelVersion).filter(ModelVersion.id == job.model_version_id).first()
    )

    if not model_version:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="模型版本不存在"
        )

    model_info = {
        "id": model_version.id,
        "name": model_version.name,
        "version": model_version.version,
        "algorithm": model_version.algorithm,
    }

    # 查询电池信息
    job_batteries = (
        db.query(TestJobBattery, BatteryUnit)
        .join(BatteryUnit, TestJobBattery.battery_id == BatteryUnit.id)
        .filter(TestJobBattery.test_job_id == job_id)
        .all()
    )

    batteries = [
        {
            "battery_id": battery.id,
            "battery_code": battery.battery_code,
            "total_cycles": battery.total_cycles,
        }
        for _, battery in job_batteries
    ]

    return TestJobDetailResponse(
        job=TestJobResponse.model_validate(job),
        model_info=model_info,
        batteries=batteries,
    )


@router.get("/jobs", response_model=list[TestJobResponse])
async def list_test_jobs(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Session = Depends(get_db),
    status_filter: Optional[str] = None,
    limit: int = 20,
    offset: int = 0,
):
    """
    列出用户的测试任务

    - 支持按状态筛选
    - 分页查询
    """
    query = db.query(TestJob).filter(TestJob.user_id == current_user.id)

    if status_filter:
        query = query.filter(TestJob.status == status_filter)

    jobs = query.order_by(TestJob.created_at.desc()).offset(offset).limit(limit).all()

    return [TestJobResponse.model_validate(job) for job in jobs]


@router.get("/jobs/{job_id}/metrics")
async def get_test_metrics(
    job_id: int,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Session = Depends(get_db),
):
    """
    获取测试指标

    返回整体指标和各电池指标
    """
    from src.models import TestJobBatteryMetric, TestJobMetricOverall

    # 验证任务所有权
    job = (
        db.query(TestJob)
        .filter(TestJob.id == job_id, TestJob.user_id == current_user.id)
        .first()
    )

    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="测试任务不存在"
        )

    # 查询整体指标
    overall_metrics = (
        db.query(TestJobMetricOverall)
        .filter(TestJobMetricOverall.test_job_id == job_id)
        .all()
    )

    # 查询各电池指标
    battery_metrics = (
        db.query(TestJobBatteryMetric)
        .filter(TestJobBatteryMetric.test_job_id == job_id)
        .all()
    )

    return {
        "job_id": job_id,
        "overall_metrics": [
            {"target": m.target, "metrics": m.metrics} for m in overall_metrics
        ],
        "battery_metrics": [
            {"battery_id": m.battery_id, "target": m.target, "metrics": m.metrics}
            for m in battery_metrics
        ],
    }


@router.get("/jobs/{job_id}/predictions")
async def get_test_predictions(
    job_id: int,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Session = Depends(get_db),
    battery_id: Optional[int] = None,
):
    """
    获取测试预测结果

    支持按电池ID筛选
    """
    from src.models import TestJobPrediction

    # 验证任务所有权
    job = (
        db.query(TestJob)
        .filter(TestJob.id == job_id, TestJob.user_id == current_user.id)
        .first()
    )

    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="测试任务不存在"
        )

    # 查询预测结果
    query = db.query(TestJobPrediction).filter(TestJobPrediction.test_job_id == job_id)

    if battery_id:
        query = query.filter(TestJobPrediction.battery_id == battery_id)

    predictions = query.order_by(
        TestJobPrediction.battery_id, TestJobPrediction.cycle_num
    ).all()

    return {
        "job_id": job_id,
        "predictions": [
            {
                "battery_id": p.battery_id,
                "cycle_num": p.cycle_num,
                "target": p.target,
                "y_true": p.y_true,
                "y_pred": p.y_pred,
            }
            for p in predictions
        ],
    }


@router.get("/jobs/{job_id}/logs")
async def get_test_logs(
    job_id: int,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Session = Depends(get_db),
    level: Optional[str] = None,
    limit: int = 1000,
):
    """
    获取测试日志（从日志文件读取）

    支持按日志级别筛选，返回最新的 limit 条日志
    """

    from src.config import settings
    from src.models import TestJobLog

    # 验证任务所有权
    job = (
        db.query(TestJob)
        .filter(TestJob.id == job_id, TestJob.user_id == current_user.id)
        .first()
    )

    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="测试任务不存在"
        )

    # 查询日志文件路径
    log_record = db.query(TestJobLog).filter(TestJobLog.test_job_id == job_id).first()

    if not log_record:
        return {
            "job_id": job_id,
            "log_file_path": None,
            "logs": [],
            "message": "日志文件尚未创建",
        }

    # 读取日志文件
    log_file_path = settings.BASE_DIR / log_record.log_file_path

    if not log_file_path.exists():
        return {
            "job_id": job_id,
            "log_file_path": str(log_record.log_file_path),
            "logs": [],
            "message": "日志文件不存在",
        }

    try:
        # 读取日志文件的最后 limit 行
        with open(str(log_file_path), "r", encoding="utf-8") as f:
            lines = f.readlines()

        # 解析日志行（格式：2026-01-13 00:45:23 - [INFO] - message）
        logs = []
        for line in lines[-limit:]:
            line = line.strip()
            if not line:
                continue

            # 简单解析
            parts = line.split(" - ", 2)
            if len(parts) >= 3:
                timestamp_str = parts[0]
                level_str = parts[1].strip("[]")
                message = parts[2]

                # 按级别过滤
                if level and level_str != level:
                    continue

                logs.append(
                    {
                        "timestamp": timestamp_str,
                        "level": level_str,
                        "message": message,
                    }
                )
            else:
                # 无法解析的行
                logs.append(
                    {
                        "timestamp": "",
                        "level": "INFO",
                        "message": line,
                    }
                )

        return {
            "job_id": job_id,
            "log_file_path": str(log_record.log_file_path),
            "total_lines": len(lines),
            "returned_lines": len(logs),
            "logs": logs,
        }

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"读取日志文件失败: {str(e)}",
        )


@router.post("/jobs/{job_id}/export")
async def export_test_results(
    job_id: int,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Session = Depends(get_db),
    format: str = Query(default="CSV", pattern="^(CSV|XLSX)$"),
):
    """
    导出测试结果

    支持CSV和XLSX格式，直接返回文件
    """
    import csv
    import io

    from src.models import BatteryUnit, TestJobMetricOverall, TestJobPrediction

    # 验证任务所有权
    job = (
        db.query(TestJob)
        .filter(TestJob.id == job_id, TestJob.user_id == current_user.id)
        .first()
    )

    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="测试任务不存在"
        )

    if bool(job.status != "SUCCEEDED"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="只能导出已完成的测试任务"
        )

    # 获取预测结果
    predictions = (
        db.query(TestJobPrediction)
        .filter(TestJobPrediction.test_job_id == job_id)
        .order_by(TestJobPrediction.battery_id, TestJobPrediction.cycle_num)
        .all()
    )

    # 获取整体指标
    overall_metrics = (
        db.query(TestJobMetricOverall)
        .filter(TestJobMetricOverall.test_job_id == job_id)
        .all()
    )

    def _format_float(value: Optional[float]) -> str:
        if value is None:
            return ""
        return f"{value:.6f}"

    def _calculate_error(
        y_true: Optional[float], y_pred: Optional[float]
    ) -> Optional[float]:
        if y_true is None or y_pred is None:
            return None
        return abs(y_true - y_pred)

    if format == "CSV":
        # 创建CSV
        output = io.StringIO()
        writer = csv.writer(output)

        # 写入元数据
        writer.writerow(["测试任务导出报告"])
        writer.writerow(["任务ID", job_id])
        writer.writerow(["目标", job.target])
        writer.writerow(["状态", job.status])
        writer.writerow(["创建时间", job.created_at.strftime("%Y-%m-%d %H:%M:%S")])
        writer.writerow([])

        # 写入整体指标
        writer.writerow(["整体指标"])
        for metric in overall_metrics:
            writer.writerow(["目标", metric.target])
            for key, value in metric.metrics.items():
                writer.writerow([key, value])
        writer.writerow([])

        # 写入预测结果表头
        writer.writerow(["预测结果"])
        writer.writerow(
            ["电池ID", "电池编号", "周期", "目标", "真实值", "预测值", "误差"]
        )

        # 写入预测结果
        for pred in predictions:
            battery = (
                db.query(BatteryUnit).filter(BatteryUnit.id == pred.battery_id).first()
            )
            pred_battery_id = cast(int, pred.battery_id)
            battery_code = (
                cast(str, battery.battery_code)
                if battery
                else f"Battery_{pred_battery_id}"
            )
            y_true = cast(Optional[float], pred.y_true)
            y_pred = cast(Optional[float], pred.y_pred)
            error = _calculate_error(y_true, y_pred)
            writer.writerow(
                [
                    pred_battery_id,
                    battery_code,
                    pred.cycle_num,
                    pred.target,
                    _format_float(y_true),
                    _format_float(y_pred),
                    _format_float(error),
                ]
            )

        # 准备响应
        output.seek(0)
        from fastapi.responses import StreamingResponse

        filename = f"test_job_{job_id}_results.csv"
        return StreamingResponse(
            io.BytesIO(
                output.getvalue().encode("utf-8-sig")
            ),  # 使用UTF-8 BOM for Excel
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    else:  # XLSX
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.worksheet import Worksheet
        except ImportError:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="XLSX导出需要安装openpyxl库",
            )

        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="无法创建Excel工作表",
            )
        ws = cast(Worksheet, ws)
        ws.title = "测试结果"

        # 标题样式
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"
        )

        row = 1

        # 写入元数据
        ws.cell(row, 1, "测试任务导出报告").font = title_font
        row += 1
        ws.cell(row, 1, "任务ID").font = header_font
        ws.cell(row, 2, job_id)
        row += 1
        ws.cell(row, 1, "目标").font = header_font
        ws.cell(row, 2, cast(str, job.target))
        row += 1
        ws.cell(row, 1, "状态").font = header_font
        ws.cell(row, 2, cast(str, job.status))
        row += 1
        ws.cell(row, 1, "创建时间").font = header_font
        ws.cell(row, 2, job.created_at.strftime("%Y-%m-%d %H:%M:%S"))
        row += 2

        # 写入整体指标
        ws.cell(row, 1, "整体指标").font = title_font
        row += 1
        for metric in overall_metrics:
            ws.cell(row, 1, "目标").font = header_font
            ws.cell(row, 2, cast(str, metric.target))
            row += 1
            for key, value in metric.metrics.items():
                ws.cell(row, 1, key)
                ws.cell(row, 2, value)
                row += 1
        row += 1

        # 写入预测结果表头
        ws.cell(row, 1, "预测结果").font = title_font
        row += 1
        headers = ["电池ID", "电池编号", "周期", "目标", "真实值", "预测值", "误差"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = header_font
            cell.fill = header_fill
        row += 1

        # 写入预测结果
        for pred in predictions:
            battery = (
                db.query(BatteryUnit).filter(BatteryUnit.id == pred.battery_id).first()
            )
            pred_battery_id = cast(int, pred.battery_id)
            battery_code = (
                cast(str, battery.battery_code)
                if battery
                else f"Battery_{pred_battery_id}"
            )
            y_true = cast(Optional[float], pred.y_true)
            y_pred = cast(Optional[float], pred.y_pred)
            error = _calculate_error(y_true, y_pred)

            ws.cell(row, 1, pred_battery_id)
            ws.cell(row, 2, battery_code)
            ws.cell(row, 3, cast(int, pred.cycle_num))
            ws.cell(row, 4, cast(str, pred.target))
            ws.cell(row, 5, y_true)
            ws.cell(row, 6, y_pred)
            ws.cell(row, 7, error)
            row += 1

        # 调整列宽
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        from fastapi.responses import StreamingResponse

        filename = f"test_job_{job_id}_results.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )


@router.websocket("/ws/jobs/{job_id}")
async def websocket_testing_progress(websocket: WebSocket, job_id: int):
    """
    WebSocket 实时测试进度

    客户端连接后会实时收到:
    - 日志消息
    - 进度更新
    - 状态变化
    """
    await manager.connect(job_id, websocket)
    try:
        while True:
            data = await websocket.receive_text()
            if data == "ping":
                await websocket.send_json({"type": "pong"})
    except WebSocketDisconnect:
        manager.disconnect(job_id, websocket)
