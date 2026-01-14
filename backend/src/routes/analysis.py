import io
from typing import Annotated, Any, cast

import pandas as pd
from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse

# 定义 Pydantic 响应模型 (与 OpenAPI 对应)
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.orm import Session

from src.models import BatteryUnit, CycleData, Dataset, User, get_db
from src.routes.auth import get_current_user

router = APIRouter()


class FeatureStat(BaseModel):
    feature_name: str
    mean: float
    variance: float
    min_val: float
    max_val: float
    corr_rul: float | None
    corr_pcl: float | None


class TrendData(BaseModel):
    cycles: list[int]
    values: list[float]


class ScatterData(BaseModel):
    points: list[list[float]]  # [[x, y], [x, y]]


class PCLDistribution(BaseModel):
    pcl_values: list[float]


class CorrelationMatrix(BaseModel):
    features: list[str]
    matrix: list[list[float]]


@router.get("/export")
async def export_analysis_report(
    dataset_id: int,
    battery_id: int,
    current_user: Annotated[User, Depends(get_current_user)],
    format: str = Query(default="PDF", pattern="^(PDF|XLSX)$"),
    db: Session = Depends(get_db),
):
    """
    导出数据分析报告

    支持PDF和XLSX格式
    """
    # 验证数据集访问权限
    dataset = db.query(Dataset).filter(Dataset.id == dataset_id).first()
    if not dataset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="数据集不存在"
        )

    # 验证电池存在
    battery = (
        db.query(BatteryUnit)
        .filter(BatteryUnit.id == battery_id, BatteryUnit.dataset_id == dataset_id)
        .first()
    )
    if not battery:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="电池不存在")

    # 获取所有分析数据
    df = get_battery_df(db, battery_id)
    stats_data = []
    feature_cols = [f"feature_{i}" for i in range(1, 9)]
    rul_series = _get_series(df, "rul") if "rul" in df.columns else None
    pcl_series = _get_series(df, "pcl") if "pcl" in df.columns else None

    for col in feature_cols:
        feature_series = _get_series(df, col)
        stats_data.append(
            {
                "feature_name": col,
                "mean": float(feature_series.mean()),
                "variance": float(feature_series.var()),
                "min_val": float(feature_series.min()),
                "max_val": float(feature_series.max()),
                "corr_rul": _safe_corr(feature_series, rul_series),
                "corr_pcl": _safe_corr(feature_series, pcl_series),
            }
        )

    if format == "XLSX":
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="XLSX导出需要安装openpyxl库",
            )

        # 创建工作簿
        wb = openpyxl.Workbook()

        # 删除默认sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # 1. 创建统计数据sheet
        ws_stats = wb.create_sheet("统计概览")
        _write_stats_sheet(ws_stats, battery, stats_data)

        # 2. 创建趋势数据sheet
        ws_trend = wb.create_sheet("特征趋势")
        _write_trend_sheet(ws_trend, db, battery_id)

        # 3. 创建相关性分析sheet
        ws_corr = wb.create_sheet("相关性分析")
        _write_correlation_sheet(ws_corr, db, battery_id)

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"analysis_report_battery_{battery.battery_code}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    else:  # PDF
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail="PDF导出功能尚未实现，请使用XLSX格式",
        )


def _write_stats_sheet(ws, battery, stats_data):
    """写入统计数据sheet"""
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    # 标题
    ws.cell(1, 1, "电池数据统计分析报告").font = Font(bold=True, size=14)
    ws.cell(2, 1, f"电池编号: {battery.battery_code}").font = Font(bold=True)
    ws.cell(3, 1, f"总周期数: {battery.total_cycles}")

    # 表头
    row = 5
    headers = ["特征", "均值", "方差", "最小值", "最大值", "与RUL相关性", "与PCL相关性"]
    header_fill = PatternFill(
        start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.font = Font(bold=True)
        cell.fill = header_fill

    # 数据
    row += 1
    for stat in stats_data:
        ws.cell(row, 1, stat["feature_name"])
        ws.cell(row, 2, stat["mean"])
        ws.cell(row, 3, stat["variance"])
        ws.cell(row, 4, stat["min_val"])
        ws.cell(row, 5, stat["max_val"])
        ws.cell(row, 6, stat["corr_rul"] if stat["corr_rul"] else "N/A")
        ws.cell(row, 7, stat["corr_pcl"] if stat["corr_pcl"] else "N/A")
        row += 1

    # 调整列宽
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 15


def _write_trend_sheet(ws, db, battery_id):
    """写入趋势数据sheet"""
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    # 获取周期数据
    cycles = (
        db.query(CycleData)
        .filter(CycleData.battery_id == battery_id)
        .order_by(CycleData.cycle_num)
        .all()
    )

    # 标题
    ws.cell(1, 1, "特征趋势数据").font = Font(bold=True, size=14)

    # 表头
    row = 3
    headers = ["周期"] + [f"Feature_{i}" for i in range(1, 9)] + ["RUL", "PCL"]
    for col, header in enumerate(headers, 1):
        ws.cell(row, col, header).font = Font(bold=True)

    # 数据
    row += 1
    for cycle in cycles:
        ws.cell(row, 1, cycle.cycle_num)
        for i in range(1, 9):
            feature_val = getattr(cycle, f"feature_{i}", None)
            ws.cell(row, i + 1, feature_val if feature_val is not None else "")
        ws.cell(row, 10, cycle.rul if cycle.rul is not None else "")
        ws.cell(row, 11, cycle.pcl if cycle.pcl is not None else "")
        row += 1

    # 调整列宽
    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 12


def _write_correlation_sheet(ws, db, battery_id):
    """写入相关性分析sheet"""
    import numpy as np
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    # 获取数据
    cycles = (
        db.query(CycleData)
        .filter(CycleData.battery_id == battery_id)
        .order_by(CycleData.cycle_num)
        .all()
    )

    if not cycles:
        ws.cell(1, 1, "无数据")
        return

    # 构建特征矩阵
    features_data = []
    for cycle in cycles:
        row_data = [getattr(cycle, f"feature_{i}", 0) or 0 for i in range(1, 9)]
        if cycle.rul is not None:
            row_data.append(cycle.rul)
        if cycle.pcl is not None:
            row_data.append(cycle.pcl)
        if len(row_data) == 10:  # 8 features + RUL + PCL
            features_data.append(row_data)

    if not features_data:
        ws.cell(1, 1, "无完整数据")
        return

    # 计算相关性矩阵
    data_array = np.array(features_data)
    corr_matrix = np.corrcoef(data_array.T)

    # 标题
    ws.cell(1, 1, "相关性矩阵").font = Font(bold=True, size=14)

    # 特征名称
    feature_names = [f"F{i}" for i in range(1, 9)] + ["RUL", "PCL"]

    # 表头（行）
    row = 3
    for col, name in enumerate(feature_names, 2):
        ws.cell(row, col, name).font = Font(bold=True)

    # 表头（列）和数据
    for i, name in enumerate(feature_names):
        row += 1
        ws.cell(row, 1, name).font = Font(bold=True)
        for j in range(len(feature_names)):
            value = corr_matrix[i, j]
            cell = ws.cell(row, j + 2, round(value, 4))

            # 根据相关性强度设置颜色
            if abs(value) > 0.7:
                cell.fill = PatternFill(
                    start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"
                )
            elif abs(value) > 0.4:
                cell.fill = PatternFill(
                    start_color="FFE66D", end_color="FFE66D", fill_type="solid"
                )

    # 调整列宽
    for col in range(1, len(feature_names) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 10


def _get_series(df: pd.DataFrame, column: str) -> pd.Series:
    return cast(pd.Series, df[column])


def _safe_corr(series: pd.Series, other: pd.Series | None) -> float:
    if other is None or other.isnull().all():
        return 0.0
    try:
        return float(series.corr(other))
    except (ValueError, TypeError):
        return 0.0


# 辅助函数：将电池数据加载为 DataFrame
def get_battery_df(db: Session, battery_id: int) -> pd.DataFrame:
    # 1. 检查电池是否存在 (且未删除)
    stmt = (
        select(CycleData)
        .where(CycleData.battery_id == battery_id, CycleData.deleted_at.is_(None))
        .order_by(CycleData.cycle_num)
    )

    results = db.execute(stmt).scalars().all()

    if not results:
        raise HTTPException(
            status_code=404, detail="No cycle data found for this battery"
        )

    # 2. 转换为 Pandas DataFrame
    data = [
        {
            "cycle_num": r.cycle_num,
            "feature_1": r.feature_1,
            "feature_2": r.feature_2,
            "feature_3": r.feature_3,
            "feature_4": r.feature_4,
            "feature_5": r.feature_5,
            "feature_6": r.feature_6,
            "feature_7": r.feature_7,
            "feature_8": r.feature_8,
            "pcl": r.pcl,
            "rul": r.rul,
        }
        for r in results
    ]
    return pd.DataFrame(data)


# --- 接口实现 ---


# 1. 获取 8x6 统计表格
@router.get("/{battery_id}/stats", response_model=list[FeatureStat])
def get_battery_stats(
    battery_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    df = get_battery_df(db, battery_id)

    stats_list = []
    feature_cols = [f"feature_{i}" for i in range(1, 9)]
    rul_series = _get_series(df, "rul") if "rul" in df.columns else None
    pcl_series = _get_series(df, "pcl") if "pcl" in df.columns else None

    for col in feature_cols:
        feature_series = _get_series(df, col)
        # 计算基础统计量
        mean_val = float(feature_series.mean())
        var_val = float(feature_series.var())
        min_val = float(feature_series.min())
        max_val = float(feature_series.max())

        # 计算相关性 (注意处理空值)
        corr_rul = _safe_corr(feature_series, rul_series)
        corr_pcl = _safe_corr(feature_series, pcl_series)

        stats_list.append(
            FeatureStat(
                feature_name=col,
                mean=round(mean_val, 4),
                variance=round(var_val, 4),
                min_val=round(min_val, 4),
                max_val=round(max_val, 4),
                corr_rul=round(corr_rul, 4),
                corr_pcl=round(corr_pcl, 4),
            )
        )

    return stats_list


# 2. 获取趋势图数据
@router.get("/{battery_id}/trend", response_model=TrendData)
def get_trend_data(
    battery_id: int,
    feature_name: str = Query(..., pattern=r"^feature_[1-8]$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # 优化：只查需要的列，不加载整个 DataFrame
    results = (
        db.query(CycleData.cycle_num, getattr(CycleData, feature_name))
        .filter(CycleData.battery_id == battery_id, CycleData.deleted_at.is_(None))
        .order_by(CycleData.cycle_num)
        .all()
    )

    if not results:
        return TrendData(cycles=[], values=[])

    # 简单的降采样策略：如果数据点超过 2000，每隔 N 个点取一个
    total_points = len(results)
    step = 1
    if total_points > 2000:
        step = total_points // 2000

    sampled_results = results[::step]

    return TrendData(
        cycles=[r[0] for r in sampled_results], values=[r[1] for r in sampled_results]
    )


# 3. 获取散点图数据 (RUL vs Feature)
@router.get("/{battery_id}/scatter", response_model=ScatterData)
def get_scatter_data(
    battery_id: int,
    feature_name: str = Query("feature_1", pattern=r"^feature_[1-8]$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    results = (
        db.query(getattr(CycleData, feature_name), CycleData.rul)
        .filter(
            CycleData.battery_id == battery_id,
            CycleData.deleted_at.is_(None),
            CycleData.rul.isnot(None),  # 过滤掉 RUL 为空的点
        )
        .all()
    )

    # ECharts 格式: [[x1, y1], [x2, y2]]
    points = [[r[0], r[1]] for r in results]

    # 同样可以做降采样
    if len(points) > 2000:
        step = len(points) // 2000
        points = points[::step]

    return ScatterData(points=points)


# 4. 获取 PCL 分布数据
@router.get("/{battery_id}/pcl-distribution", response_model=PCLDistribution)
def get_pcl_distribution(
    battery_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    results = (
        db.query(CycleData.pcl)
        .filter(CycleData.battery_id == battery_id, CycleData.deleted_at.is_(None))
        .all()
    )

    return PCLDistribution(pcl_values=[r[0] for r in results if r[0] is not None])


# 5. 获取相关性矩阵 (热力图)
@router.get("/{battery_id}/correlation-matrix", response_model=CorrelationMatrix)
def get_correlation_matrix(
    battery_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    df = get_battery_df(db, battery_id)
    feature_cols = [f"feature_{i}" for i in range(1, 9)]

    # 计算相关矩阵
    feature_df: pd.DataFrame = df.loc[:, feature_cols]
    corr_matrix = feature_df.corr()

    # 填充 NaN (如果有常数列，相关性会是 NaN) 为 0
    corr_matrix = corr_matrix.fillna(0)

    # 转换为 8x8 二维数组
    matrix_data = corr_matrix.values.tolist()

    return CorrelationMatrix(features=feature_cols, matrix=matrix_data)


# 6. 导出分析报告
@router.get("/{battery_id}/export-report")
def export_battery_report(
    battery_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    导出电池分析报告 (Excel)
    包含:
    - Sheet1: 原始数据 (Raw Data)
    - Sheet2: 特征统计 (Statistics)
    - Sheet3: 相关性矩阵 (Correlation)
    """
    df = get_battery_df(db, battery_id)

    # 1. 准备统计数据
    stats_data = []
    feature_cols = [f"feature_{i}" for i in range(1, 9)]
    rul_series = _get_series(df, "rul") if "rul" in df.columns else None
    pcl_series = _get_series(df, "pcl") if "pcl" in df.columns else None

    for col in feature_cols:
        feature_series = _get_series(df, col)
        corr_rul = _safe_corr(feature_series, rul_series)
        corr_pcl = _safe_corr(feature_series, pcl_series)

        stats_data.append(
            {
                "Feature": col,
                "Mean": float(feature_series.mean()),
                "Variance": float(feature_series.var()),
                "Min": float(feature_series.min()),
                "Max": float(feature_series.max()),
                "Corr with RUL": corr_rul,
                "Corr with PCL": corr_pcl,
            }
        )
    df_stats = pd.DataFrame(stats_data)

    # 2. 准备相关性矩阵
    feature_df: pd.DataFrame = df.loc[:, feature_cols]
    df_corr = feature_df.corr().fillna(0)

    # 3. 写入 Excel
    output = io.BytesIO()
    with pd.ExcelWriter(cast(Any, output), engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Raw Data", index=False)
        df_stats.to_excel(writer, sheet_name="Statistics", index=False)
        df_corr.to_excel(writer, sheet_name="Correlation", index=True)

    output.seek(0)

    filename = f"battery_{battery_id}_analysis_report.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
